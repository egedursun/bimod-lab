from io import BytesIO

from docx import Document as ms_dx
from pptx import Presentation as ms_px
from openpyxl import load_workbook as ms_ex


def decode_docx(file_bytes):
    file_stream = BytesIO(file_bytes)
    document = ms_dx(file_stream)
    return '\n'.join([paragraph.text for paragraph in document.paragraphs])


def decode_pptx(file_bytes):
    file_stream = BytesIO(file_bytes)
    presentation = ms_px(file_stream)
    text = []
    for slide in presentation.slides:
        for shape in slide.shapes:
            if hasattr(shape, "text"):
                text.append(shape.text)
    return '\n'.join(text)


def decode_xlsx(file_bytes):
    file_stream = BytesIO(file_bytes)
    workbook = ms_ex(file_stream, data_only=True)
    text = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text.append('\t'.join([str(cell) if cell is not None else '' for cell in row]))
    return '\n'.join(text)
